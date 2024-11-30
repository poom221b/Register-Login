from openpyxl import Workbook, load_workbook
import bcrypt

def is_valid_password(password): #เช็คพาสเวิดให้มันเกิน 8 ตัว ละก็ต้องมีเลข
    has_number = any(char.isdigit() for char in password)
    has_letter = any(char.isalpha() for char in password)
    return has_number and has_letter and len(password) >= 8

def register():
    try:
        workbook = load_workbook("users.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Username", "Password"])

    username = input("Enter your username: ")

    # ตรวจสอบว่า username มีอยู่แล้วมั้ย
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if username == row[0]:
            print("Username already exists! 😅 Please try again.")
            return
    #ถ้ายังไม่มีคนใช้ก็ไปต่อ
    while True:
        password = input("Enter your password: ")
        if is_valid_password(password):
            print("Password is valid! 🎉")
            break
        else:
            print("Invalid password! 😢 Must contain letters, numbers, and be at least 8 characters long.")
    #เข้ารหัสpassword
    hashed_password = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

    sheet.append([username, hashed_password])
    workbook.save("users.xlsx")
    print("Registration successful! 🎉")


def login():
    try:
        workbook = load_workbook("users.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        print("No users found! Please register first. 😢")
        return

    username = input("Enter your username: ")
    password = input("Enter your password: ")

    # ตรวจสอบข้อมูลว่าตรงกับในไฟล์ไหม loop ผ่านทุกแถวในเอกเซล ยกเว้นแถวแรก เพราะแถวแรกเอาไว้ใส่ชื่อ ละถ้าเจอที่ยูเซอเนมกะพาสตรงกันก็รีเทินซักเซส
    for row in sheet.iter_rows(min_row=2, values_only=True):
        db_username, db_password = row
        if username == db_username and bcrypt.checkpw(password.encode(), db_password if isinstance(db_password,
                                                                                                   bytes) else db_password.encode()):
            print("Login successful! ✨")
            return
    print("Invalid username or password! 😞")


def main():
    while True:
        print("\n--- Welcome ---")
        print("1. Register")
        print("2. Login")
        print("3. Exit")

        choice = input("Enter your choice: ")
        if choice == "1":
            register()
        elif choice == "2":
            login()
        elif choice == "3":
            print("Goodbye! 👋")
            break
        else:
            print("Invalid choice, please try again! 😅")


if __name__ == "__main__":
    main()
